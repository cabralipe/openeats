from django.db import models
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
import csv
import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from rest_framework import permissions, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from merenda_semed.authentication import QueryParamJWTAuthentication

from .models import (
    Delivery,
    DeliveryItem,
    Notification,
    Responsible,
    SchoolStockBalance,
    Supplier,
    SupplierReceipt,
    Supply,
    StockBalance,
    StockMovement,
)
from .serializers import (
    DeliverySerializer,
    NotificationSerializer,
    ResponsibleSerializer,
    SchoolStockBalanceSerializer,
    SchoolStockBulkLimitItemSerializer,
    SchoolStockLimitUpdateSerializer,
    SupplierReceiptConferenceInputSerializer,
    SupplierReceiptSerializer,
    SupplierSerializer,
    SupplySerializer,
    StockBalanceSerializer,
    StockMovementSerializer,
)



def _pdf_text(value):
    text = '' if value is None else str(value)
    return text.encode('latin-1', 'replace').decode('latin-1')


def _format_filter_value(value):
    if value in (None, ''):
        return 'Todos'
    return str(value)


def _start_pdf_page(pdf, title, generated_at, filters, page_number):
    width, height = A4
    left = 32
    right = width - 32

    pdf.setFillColor(colors.HexColor('#1f3a6d'))
    pdf.rect(0, height - 84, width, 84, stroke=0, fill=1)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(left, height - 32, _pdf_text('SEMED - Prestacao de Contas'))
    pdf.setFont('Helvetica', 11)
    pdf.drawString(left, height - 50, _pdf_text(title))
    pdf.setFont('Helvetica', 9)
    pdf.drawRightString(right, height - 32, _pdf_text(f"Pagina {page_number}"))
    pdf.drawRightString(right, height - 50, _pdf_text(f"Gerado em: {generated_at.strftime('%Y-%m-%d %H:%M')}"))

    y = height - 102
    pdf.setFillColor(colors.HexColor('#f4f6fb'))
    pdf.rect(left, y - 28, right - left, 28, stroke=0, fill=1)
    pdf.setFillColor(colors.HexColor('#1b2a41'))
    pdf.setFont('Helvetica', 8.5)
    filters_text = ' | '.join([f"{label}: {_format_filter_value(value)}" for label, value in filters])
    pdf.drawString(left + 8, y - 17, _pdf_text(filters_text))
    return y - 40


def _draw_pdf_footer(pdf, page_number):
    width, _ = A4
    pdf.setStrokeColor(colors.HexColor('#d2d8e6'))
    pdf.line(32, 26, width - 32, 26)
    pdf.setFont('Helvetica', 8)
    pdf.setFillColor(colors.HexColor('#5b6578'))
    pdf.drawString(32, 14, _pdf_text('Documento de prestacao de contas - uso interno SEMED'))
    pdf.drawRightString(width - 32, 14, _pdf_text(f"Pagina {page_number}"))


class SupplyViewSet(viewsets.ModelViewSet):
    queryset = Supply.objects.all().order_by('name')
    serializer_class = SupplySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.query_params.get('q')
        category = self.request.query_params.get('category')
        is_active = self.request.query_params.get('is_active')
        if query:
            queryset = queryset.filter(name__icontains=query)
        if category:
            queryset = queryset.filter(category__icontains=category)
        if is_active in ['true', 'false']:
            queryset = queryset.filter(is_active=is_active == 'true')
        return queryset

    def perform_create(self, serializer):
        supply = serializer.save()
        StockBalance.objects.get_or_create(supply=supply)


class StockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StockBalance.objects.select_related('supply').all().order_by('supply__name')
    serializer_class = StockBalanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.query_params.get('q')
        category = self.request.query_params.get('category')
        low_stock = self.request.query_params.get('low_stock')
        if query:
            queryset = queryset.filter(supply__name__icontains=query)
        if category:
            queryset = queryset.filter(supply__category__icontains=category)
        if low_stock in ['true', 'false']:
            if low_stock == 'true':
                queryset = queryset.filter(quantity__lt=models.F('supply__min_stock'))
            else:
                queryset = queryset.filter(quantity__gte=models.F('supply__min_stock'))
        return queryset


class StockMovementViewSet(viewsets.ModelViewSet):
    queryset = StockMovement.objects.select_related('supply').all().order_by('-created_at')
    serializer_class = StockMovementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        movement_type = self.request.query_params.get('type')
        supply = self.request.query_params.get('supply')
        school = self.request.query_params.get('school')
        if date_from:
            queryset = queryset.filter(movement_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(movement_date__lte=date_to)
        if movement_type:
            queryset = queryset.filter(type=movement_type)
        if supply:
            queryset = queryset.filter(supply_id=supply)
        if school:
            queryset = queryset.filter(school_id=school)
        return queryset


class ResponsibleViewSet(viewsets.ModelViewSet):
    queryset = Responsible.objects.all().order_by('name')
    serializer_class = ResponsibleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.query_params.get('q')
        position = self.request.query_params.get('position')
        is_active = self.request.query_params.get('is_active')
        if query:
            queryset = queryset.filter(name__icontains=query)
        if position:
            queryset = queryset.filter(position__icontains=position)
        if is_active in ['true', 'false']:
            queryset = queryset.filter(is_active=is_active == 'true')
        return queryset


class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all().order_by('name')
    serializer_class = SupplierSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.query_params.get('q')
        is_active = self.request.query_params.get('is_active')
        if query:
            queryset = queryset.filter(models.Q(name__icontains=query) | models.Q(document__icontains=query))
        if is_active in ['true', 'false']:
            queryset = queryset.filter(is_active=is_active == 'true')
        return queryset


class SupplierReceiptViewSet(viewsets.ModelViewSet):
    queryset = SupplierReceipt.objects.select_related('supplier', 'school', 'created_by').prefetch_related('items__supply').all().order_by('-expected_date', '-created_at')
    serializer_class = SupplierReceiptSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        supplier = self.request.query_params.get('supplier')
        school = self.request.query_params.get('school')
        status_value = self.request.query_params.get('status')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if supplier:
            queryset = queryset.filter(supplier_id=supplier)
        if school:
            queryset = queryset.filter(school_id=school)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if date_from:
            queryset = queryset.filter(expected_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(expected_date__lte=date_to)
        return queryset

    @action(detail=True, methods=['post'])
    def start_conference(self, request, pk=None):
        receipt = self.get_object()
        if receipt.status in [SupplierReceipt.Status.CONFERRED, SupplierReceipt.Status.CANCELLED]:
            raise ValidationError('Nao e possivel iniciar conferencia para este recebimento.')
        if receipt.status != SupplierReceipt.Status.IN_CONFERENCE:
            receipt.status = SupplierReceipt.Status.IN_CONFERENCE
            receipt.conference_started_at = timezone.now()
            receipt.save(update_fields=['status', 'conference_started_at', 'updated_at'])
        return Response(self.get_serializer(receipt).data)

    @action(detail=True, methods=['post'])
    def submit_conference(self, request, pk=None):
        payload = SupplierReceiptConferenceInputSerializer(data=request.data)
        payload.is_valid(raise_exception=True)

        with transaction.atomic():
            receipt = SupplierReceipt.objects.select_for_update().get(id=pk)
            if receipt.status in [SupplierReceipt.Status.CONFERRED, SupplierReceipt.Status.CANCELLED]:
                raise ValidationError('Nao e possivel concluir conferencia para este recebimento.')

            receipt_items = {
                str(item.id): item
                for item in receipt.items.select_for_update()
            }

            payload_items = payload.validated_data['items']
            payload_ids = {str(entry['item_id']) for entry in payload_items}
            missing_ids = set(receipt_items.keys()) - payload_ids
            if missing_ids:
                raise ValidationError('Envie a conferencia de todos os itens do recebimento.')

            for entry in payload_items:
                item = receipt_items.get(str(entry['item_id']))
                if not item:
                    raise ValidationError('Item de conferencia nao pertence ao recebimento.')
                resolved_supply = item.supply
                if not resolved_supply:
                    raw_name = (item.raw_name or '').strip()
                    category = (item.category or '').strip()
                    if not raw_name:
                        raise ValidationError('Item novo sem nome para criar insumo automaticamente.')
                    if not category:
                        raise ValidationError(f'Categoria obrigatoria para criar o item "{raw_name}".')

                    resolved_supply = Supply.objects.filter(
                        name__iexact=raw_name,
                        category__iexact=category,
                        unit=item.unit,
                    ).first()
                    if not resolved_supply:
                        resolved_supply = Supply.objects.create(
                            name=raw_name,
                            category=category,
                            unit=item.unit,
                            min_stock=0,
                            is_active=True,
                        )
                    item.supply_created = resolved_supply

                quantity = entry['received_quantity']
                item.received_quantity = quantity
                item.divergence_note = entry.get('note', '')
                item.save(update_fields=['received_quantity', 'divergence_note', 'supply_created'])

                if receipt.school_id:
                    school_balance, _ = SchoolStockBalance.objects.select_for_update().get_or_create(
                        school=receipt.school,
                        supply=resolved_supply,
                        defaults={'quantity': 0, 'min_stock': 0},
                    )
                    school_balance.quantity += quantity
                    school_balance.save()
                else:
                    balance, _ = StockBalance.objects.select_for_update().get_or_create(
                        supply=resolved_supply,
                        defaults={'quantity': 0},
                    )
                    balance.quantity += quantity
                    balance.save(update_fields=['quantity'])

                StockMovement.objects.create(
                    supply=resolved_supply,
                    school=receipt.school if receipt.school_id else None,
                    type=StockMovement.Types.IN,
                    quantity=quantity,
                    movement_date=receipt.expected_date,
                    note=f'Entrada por recebimento de fornecedor {receipt.id}.',
                    created_by=request.user,
                )

            now = timezone.now()
            receipt.status = SupplierReceipt.Status.CONFERRED
            receipt.conference_started_at = receipt.conference_started_at or now
            receipt.conference_finished_at = now
            receipt.sender_signature = payload.validated_data['sender_signature_data']
            receipt.sender_signed_by = payload.validated_data['sender_signer_name']
            receipt.receiver_signature = payload.validated_data['receiver_signature_data']
            receipt.receiver_signed_by = payload.validated_data['receiver_signer_name']
            receipt.save(update_fields=[
                'status',
                'conference_started_at',
                'conference_finished_at',
                'sender_signature',
                'sender_signed_by',
                'receiver_signature',
                'receiver_signed_by',
                'updated_at',
            ])

        receipt = self.get_object()
        return Response(self.get_serializer(receipt).data)


class StockExportCsvView(viewsets.ViewSet):

    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = StockBalance.objects.select_related('supply').all().order_by('supply__name')
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=\"stock.csv\"'
        writer = csv.writer(response)
        writer.writerow(['Insumo', 'Categoria', 'Unidade', 'Quantidade', 'Minimo', 'Status', 'Diferenca'])
        for balance in queryset:
            diff = balance.quantity - balance.supply.min_stock
            if balance.quantity < balance.supply.min_stock:
                status = 'BAIXO'
            elif balance.quantity >= balance.supply.min_stock * 2:
                status = 'ALTO'
            else:
                status = 'NORMAL'
            writer.writerow([
                balance.supply.name,
                balance.supply.category,
                balance.supply.unit,
                balance.quantity,
                balance.supply.min_stock,
                status,
                diff,
            ])
        return response


class StockExportPdfView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = StockBalance.objects.select_related('supply').all().order_by('supply__category', 'supply__name')

        items = []
        for balance in queryset:
            diff = float(balance.quantity) - float(balance.supply.min_stock)
            if balance.quantity < balance.supply.min_stock:
                status = 'BAIXO'
            elif balance.quantity >= balance.supply.min_stock * 2:
                status = 'ALTO'
            else:
                status = 'NORMAL'
            items.append({
                'name': balance.supply.name,
                'category': balance.supply.category or 'Sem Categoria',
                'unit': balance.supply.unit,
                'quantity': float(balance.quantity),
                'min_stock': float(balance.supply.min_stock),
                'status': status,
                'diff': diff,
            })

        low_total = sum(1 for entry in items if entry['status'] == 'BAIXO')
        normal_total = sum(1 for entry in items if entry['status'] == 'NORMAL')
        high_total = sum(1 for entry in items if entry['status'] == 'ALTO')
        total_items = len(items)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=\"stock_report.pdf\"'
        pdf = canvas.Canvas(response, pagesize=A4)
        _, height = A4
        generated_at = timezone.now()
        page_number = 1
        filters = [('Tipo', 'Relatorio geral de estoque')]
        y = _start_pdf_page(pdf, 'Relatorio de Estoque', generated_at, filters, page_number)

        pdf.setFont('Helvetica-Bold', 10)
        pdf.setFillColor(colors.HexColor('#1b2a41'))
        pdf.drawString(32, y, _pdf_text('Resumo Executivo'))
        y -= 14
        pdf.setFillColor(colors.HexColor('#f7f9fc'))
        pdf.rect(32, y - 42, 270, 42, stroke=1, fill=1)
        pdf.setFillColor(colors.HexColor('#1b2a41'))
        pdf.setFont('Helvetica', 9)
        pdf.drawString(40, y - 14, _pdf_text(f"Total de insumos analisados: {total_items}"))
        pdf.drawString(40, y - 28, _pdf_text(f"Baixo: {low_total} | Normal: {normal_total} | Alto: {high_total}"))
        y -= 56

        def draw_table_header(y_pos):
            pdf.setFillColor(colors.HexColor('#e3e9f5'))
            pdf.rect(32, y_pos - 14, 530, 16, stroke=0, fill=1)
            pdf.setFillColor(colors.HexColor('#1b2a41'))
            pdf.setFont('Helvetica-Bold', 8.5)
            pdf.drawString(36, y_pos - 10, _pdf_text('INSUMO'))
            pdf.drawString(205, y_pos - 10, _pdf_text('CATEGORIA'))
            pdf.drawString(315, y_pos - 10, _pdf_text('UNID'))
            pdf.drawRightString(392, y_pos - 10, _pdf_text('QTD'))
            pdf.drawRightString(452, y_pos - 10, _pdf_text('MIN'))
            pdf.drawRightString(512, y_pos - 10, _pdf_text('DIF'))
            pdf.drawRightString(556, y_pos - 10, _pdf_text('STATUS'))
            return y_pos - 20

        y = draw_table_header(y)
        current_category = None
        for entry in items:
            if y < 64:
                _draw_pdf_footer(pdf, page_number)
                pdf.showPage()
                page_number += 1
                y = _start_pdf_page(pdf, 'Relatorio de Estoque', generated_at, filters, page_number)
                y = draw_table_header(y)
                current_category = None

            if entry['category'] != current_category:
                current_category = entry['category']
                pdf.setFillColor(colors.HexColor('#f0f3fa'))
                pdf.rect(32, y - 12, 530, 14, stroke=0, fill=1)
                pdf.setFillColor(colors.HexColor('#344563'))
                pdf.setFont('Helvetica-Bold', 8)
                pdf.drawString(36, y - 9, _pdf_text(current_category[:50]))
                y -= 15
                if y < 64:
                    _draw_pdf_footer(pdf, page_number)
                    pdf.showPage()
                    page_number += 1
                    y = _start_pdf_page(pdf, 'Relatorio de Estoque', generated_at, filters, page_number)
                    y = draw_table_header(y)

            pdf.setFont('Helvetica', 8)
            pdf.setFillColor(colors.black)
            pdf.drawString(36, y - 8, _pdf_text(entry['name'][:36]))
            pdf.drawString(205, y - 8, _pdf_text(entry['category'][:20]))
            pdf.drawString(315, y - 8, _pdf_text(entry['unit']))
            pdf.drawRightString(392, y - 8, _pdf_text(f"{entry['quantity']:.2f}"))
            pdf.drawRightString(452, y - 8, _pdf_text(f"{entry['min_stock']:.2f}"))
            pdf.drawRightString(512, y - 8, _pdf_text(f"{entry['diff']:.2f}"))
            pdf.drawRightString(556, y - 8, _pdf_text(entry['status']))
            pdf.setStrokeColor(colors.HexColor('#e8edf6'))
            pdf.line(32, y - 11, 562, y - 11)
            y -= 14

        _draw_pdf_footer(pdf, page_number)
        pdf.save()
        return response


class StockExportXlsxView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = StockBalance.objects.select_related('supply').all().order_by('supply__category', 'supply__name')

        workbook = Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = 'Resumo'
        summary_sheet.append(['Categoria', 'Total Itens', 'Itens Baixos', 'Itens Normais', 'Itens Altos'])
        
        # All items sheet
        items_sheet = workbook.create_sheet(title='Todos os Itens')
        items_sheet.append(['Insumo', 'Categoria', 'Unidade', 'Quantidade', 'Estoque Minimo', 'Status', 'Diferenca'])
        
        # Low stock sheet
        low_sheet = workbook.create_sheet(title='Estoque Baixo')
        low_sheet.append(['Insumo', 'Categoria', 'Unidade', 'Quantidade', 'Estoque Minimo', 'Falta'])
        
        # Normal stock sheet
        normal_sheet = workbook.create_sheet(title='Estoque Normal')
        normal_sheet.append(['Insumo', 'Categoria', 'Unidade', 'Quantidade', 'Estoque Minimo'])
        
        # High stock sheet
        high_sheet = workbook.create_sheet(title='Estoque Alto')
        high_sheet.append(['Insumo', 'Categoria', 'Unidade', 'Quantidade', 'Estoque Minimo', 'Excesso'])

        category_stats = {}
        
        for balance in queryset:
            cat = balance.supply.category or 'Sem Categoria'
            if cat not in category_stats:
                category_stats[cat] = {'total': 0, 'low': 0, 'normal': 0, 'high': 0}
            
            category_stats[cat]['total'] += 1
            diff = float(balance.quantity) - float(balance.supply.min_stock)
            
            if balance.quantity < balance.supply.min_stock:
                status = 'BAIXO'
                category_stats[cat]['low'] += 1
                low_sheet.append([
                    balance.supply.name,
                    cat,
                    balance.supply.unit,
                    float(balance.quantity),
                    float(balance.supply.min_stock),
                    abs(diff),
                ])
            elif balance.quantity >= balance.supply.min_stock * 2:
                status = 'ALTO'
                category_stats[cat]['high'] += 1
                high_sheet.append([
                    balance.supply.name,
                    cat,
                    balance.supply.unit,
                    float(balance.quantity),
                    float(balance.supply.min_stock),
                    diff,
                ])
            else:
                status = 'NORMAL'
                category_stats[cat]['normal'] += 1
                normal_sheet.append([
                    balance.supply.name,
                    cat,
                    balance.supply.unit,
                    float(balance.quantity),
                    float(balance.supply.min_stock),
                ])
            
            items_sheet.append([
                balance.supply.name,
                cat,
                balance.supply.unit,
                float(balance.quantity),
                float(balance.supply.min_stock),
                status,
                diff,
            ])
        
        for cat, stats in category_stats.items():
            summary_sheet.append([cat, stats['total'], stats['low'], stats['normal'], stats['high']])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=\"stock_report.xlsx\"'
        return response


class DeliveryExportPdfView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = Delivery.objects.select_related('school').prefetch_related('items__supply').all().order_by('-delivery_date')
        school = request.query_params.get('school')
        status_value = request.query_params.get('status')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if school:
            queryset = queryset.filter(school_id=school)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if date_from:
            queryset = queryset.filter(delivery_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(delivery_date__lte=date_to)

        deliveries = list(queryset)
        total_deliveries = len(deliveries)
        total_items = sum(delivery.items.count() for delivery in deliveries)
        draft_total = sum(1 for delivery in deliveries if delivery.status == Delivery.Status.DRAFT)
        sent_total = sum(1 for delivery in deliveries if delivery.status == Delivery.Status.SENT)
        conferred_total = sum(1 for delivery in deliveries if delivery.status == Delivery.Status.CONFERRED)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=\"deliveries.pdf\"'
        pdf = canvas.Canvas(response, pagesize=A4)
        _, height = A4
        generated_at = timezone.now()
        page_number = 1
        filters = [
            ('Escola', school),
            ('Status', status_value),
            ('Data inicial', date_from),
            ('Data final', date_to),
        ]
        y = _start_pdf_page(pdf, 'Relatorio de Entregas', generated_at, filters, page_number)

        pdf.setFont('Helvetica', 9)
        pdf.setFillColor(colors.HexColor('#1b2a41'))
        pdf.drawString(32, y, _pdf_text(
            f"Total de entregas: {total_deliveries} | Itens planejados: {total_items} | "
            f"Rascunho: {draft_total} | Enviada: {sent_total} | Conferida: {conferred_total}"
        ))
        y -= 18

        for delivery in deliveries:
            delivery_items = list(delivery.items.all())
            required_height = 78 + (len(delivery_items) * 12)
            if y < required_height + 40:
                _draw_pdf_footer(pdf, page_number)
                pdf.showPage()
                page_number += 1
                y = _start_pdf_page(pdf, 'Relatorio de Entregas', generated_at, filters, page_number)

            pdf.setFillColor(colors.HexColor('#f7f9fc'))
            pdf.rect(32, y - required_height + 8, 530, required_height, stroke=1, fill=1)
            pdf.setFillColor(colors.HexColor('#1b2a41'))
            pdf.setFont('Helvetica-Bold', 10)
            pdf.drawString(
                40,
                y - 8,
                _pdf_text(f"{delivery.delivery_date} | {delivery.school.name} | {delivery.get_status_display()} | Entrega: {delivery.id}"),
            )
            pdf.setFont('Helvetica', 8.5)
            conference_at = delivery.conference_submitted_at.strftime('%Y-%m-%d %H:%M') if delivery.conference_submitted_at else '-'
            sender_name = delivery.sender_signed_by or delivery.responsible_name or '-'
            receiver_name = delivery.receiver_signed_by or delivery.conference_signed_by or '-'
            pdf.drawString(40, y - 22, _pdf_text(f"Conferida em: {conference_at}"))
            pdf.drawString(200, y - 22, _pdf_text(f"Entregador: {sender_name[:35]}"))
            pdf.drawString(390, y - 22, _pdf_text(f"Recebedor: {receiver_name[:28]}"))

            pdf.setFont('Helvetica-Bold', 8)
            pdf.drawString(40, y - 36, _pdf_text('Item'))
            pdf.drawRightString(372, y - 36, _pdf_text('Prev.'))
            pdf.drawRightString(442, y - 36, _pdf_text('Receb.'))
            pdf.drawRightString(502, y - 36, _pdf_text('Falta'))
            pdf.drawString(510, y - 36, _pdf_text('Obs.'))
            pdf.setFont('Helvetica', 8)

            line_y = y - 48
            for item in delivery_items:
                shortage = 0.0
                if item.received_quantity is not None:
                    shortage = max(float(item.planned_quantity - item.received_quantity), 0.0)
                received = '-' if item.received_quantity is None else f"{float(item.received_quantity):.2f}"
                pdf.drawString(40, line_y, _pdf_text(item.supply.name[:42]))
                pdf.drawRightString(372, line_y, _pdf_text(f"{float(item.planned_quantity):.2f}"))
                pdf.drawRightString(442, line_y, _pdf_text(received))
                pdf.drawRightString(502, line_y, _pdf_text(f"{shortage:.2f}" if item.received_quantity is not None else '-'))
                pdf.drawString(510, line_y, _pdf_text((item.divergence_note or '-')[:10]))
                line_y -= 12

            y -= required_height + 8

        _draw_pdf_footer(pdf, page_number)
        pdf.save()
        return response


class DeliveryExportXlsxView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = Delivery.objects.select_related('school').prefetch_related('items').all().order_by('-delivery_date')
        school = request.query_params.get('school')
        status_value = request.query_params.get('status')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if school:
            queryset = queryset.filter(school_id=school)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if date_from:
            queryset = queryset.filter(delivery_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(delivery_date__lte=date_to)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = 'Entregas'
        summary_sheet.append(['Data', 'Escola', 'Status', 'Itens', 'Responsavel', 'Telefone', 'Observacoes'])

        items_sheet = workbook.create_sheet(title='Itens')
        items_sheet.append([
            'Data',
            'Escola',
            'Status',
            'Insumo',
            'Unidade',
            'Quantidade Planejada',
            'Quantidade Recebida',
            'Falta',
            'Observacao Divergencia',
        ])

        for delivery in queryset:
            items = list(delivery.items.select_related('supply').all())
            summary_sheet.append([
                str(delivery.delivery_date),
                delivery.school.name,
                delivery.get_status_display(),
                len(items),
                delivery.responsible_name,
                delivery.responsible_phone,
                delivery.notes,
            ])
            for item in items:
                shortage = None
                if item.received_quantity is not None:
                    shortage_value = item.planned_quantity - item.received_quantity
                    shortage = float(shortage_value) if shortage_value > 0 else 0
                items_sheet.append([
                    str(delivery.delivery_date),
                    delivery.school.name,
                    delivery.get_status_display(),
                    item.supply.name,
                    item.supply.unit,
                    float(item.planned_quantity),
                    float(item.received_quantity) if item.received_quantity is not None else None,
                    shortage,
                    item.divergence_note,
                ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=\"deliveries.xlsx\"'
        return response


class ConsumptionExportPdfView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = StockMovement.objects.select_related('supply', 'school').filter(type=StockMovement.Types.OUT).order_by('-movement_date')
        supply = request.query_params.get('supply')
        school = request.query_params.get('school')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if supply:
            queryset = queryset.filter(supply_id=supply)
        if school:
            queryset = queryset.filter(school_id=school)
        if date_from:
            queryset = queryset.filter(movement_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(movement_date__lte=date_to)

        movements = list(queryset)
        total_records = len(movements)
        total_quantity = sum(float(movement.quantity) for movement in movements)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=\"consumption.pdf\"'
        pdf = canvas.Canvas(response, pagesize=A4)
        _, height = A4
        generated_at = timezone.now()
        page_number = 1
        filters = [
            ('Insumo', supply),
            ('Escola', school),
            ('Data inicial', date_from),
            ('Data final', date_to),
        ]
        y = _start_pdf_page(pdf, 'Relatorio de Consumo', generated_at, filters, page_number)

        pdf.setFont('Helvetica', 9)
        pdf.setFillColor(colors.HexColor('#1b2a41'))
        pdf.drawString(
            32,
            y,
            _pdf_text(f"Total de registros: {total_records} | Quantidade total de consumo: {total_quantity:.2f}"),
        )
        y -= 18

        def draw_table_header(y_pos):
            pdf.setFillColor(colors.HexColor('#e3e9f5'))
            pdf.rect(32, y_pos - 14, 530, 16, stroke=0, fill=1)
            pdf.setFillColor(colors.HexColor('#1b2a41'))
            pdf.setFont('Helvetica-Bold', 8.5)
            pdf.drawString(36, y_pos - 10, _pdf_text('DATA'))
            pdf.drawString(90, y_pos - 10, _pdf_text('ESCOLA'))
            pdf.drawString(230, y_pos - 10, _pdf_text('INSUMO'))
            pdf.drawRightString(455, y_pos - 10, _pdf_text('QTD'))
            pdf.drawString(462, y_pos - 10, _pdf_text('UN'))
            pdf.drawString(494, y_pos - 10, _pdf_text('OBS.'))
            return y_pos - 20

        y = draw_table_header(y)
        for movement in movements:
            if y < 64:
                _draw_pdf_footer(pdf, page_number)
                pdf.showPage()
                page_number += 1
                y = _start_pdf_page(pdf, 'Relatorio de Consumo', generated_at, filters, page_number)
                y = draw_table_header(y)

            school_name = movement.school.name if movement.school else 'Central'
            pdf.setFont('Helvetica', 8)
            pdf.setFillColor(colors.black)
            pdf.drawString(36, y - 8, _pdf_text(str(movement.movement_date)))
            pdf.drawString(90, y - 8, _pdf_text(school_name[:24]))
            pdf.drawString(230, y - 8, _pdf_text(movement.supply.name[:34]))
            pdf.drawRightString(455, y - 8, _pdf_text(f"{float(movement.quantity):.2f}"))
            pdf.drawString(462, y - 8, _pdf_text(movement.supply.unit))
            pdf.drawString(494, y - 8, _pdf_text((movement.note or '-')[:12]))
            pdf.setStrokeColor(colors.HexColor('#e8edf6'))
            pdf.line(32, y - 11, 562, y - 11)
            y -= 14

        _draw_pdf_footer(pdf, page_number)
        pdf.save()
        return response


class ConsumptionExportXlsxView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = StockMovement.objects.select_related('supply').filter(type=StockMovement.Types.OUT).order_by('-movement_date')
        supply = request.query_params.get('supply')
        school = request.query_params.get('school')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if supply:
            queryset = queryset.filter(supply_id=supply)
        if school:
            queryset = queryset.filter(school_id=school)
        if date_from:
            queryset = queryset.filter(movement_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(movement_date__lte=date_to)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Consumo'
        sheet.append(['Data', 'Insumo', 'Unidade', 'Quantidade', 'Observacao'])

        totals_sheet = workbook.create_sheet(title='Resumo por Insumo')
        totals_sheet.append(['Insumo', 'Unidade', 'Quantidade Total'])

        totals = {}
        for movement in queryset:
            sheet.append([
                str(movement.movement_date),
                movement.supply.name,
                movement.supply.unit,
                float(movement.quantity),
                movement.note,
            ])
            key = movement.supply_id
            if key not in totals:
                totals[key] = {
                    'name': movement.supply.name,
                    'unit': movement.supply.unit,
                    'total': 0.0,
                }
            totals[key]['total'] += float(movement.quantity)

        for entry in totals.values():
            totals_sheet.append([entry['name'], entry['unit'], entry['total']])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=\"consumption.xlsx\"'
        return response


class SupplierReceiptExportPdfView(viewsets.ViewSet):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        queryset = SupplierReceipt.objects.select_related('supplier', 'school').prefetch_related('items__supply', 'items__supply_created').all().order_by('-expected_date', '-created_at')
        supplier = request.query_params.get('supplier')
        school = request.query_params.get('school')
        status_value = request.query_params.get('status')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if supplier:
            queryset = queryset.filter(supplier_id=supplier)
        if school:
            queryset = queryset.filter(school_id=school)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if date_from:
            queryset = queryset.filter(expected_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(expected_date__lte=date_to)

        receipts = list(queryset)
        total_receipts = len(receipts)
        total_items = sum(receipt.items.count() for receipt in receipts)
        conferred_total = sum(1 for receipt in receipts if receipt.status == SupplierReceipt.Status.CONFERRED)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=\"supplier_receipts.pdf\"'
        pdf = canvas.Canvas(response, pagesize=A4)
        _, height = A4
        generated_at = timezone.now()
        page_number = 1
        filters = [
            ('Fornecedor', supplier),
            ('Escola', school),
            ('Status', status_value),
            ('Data inicial', date_from),
            ('Data final', date_to),
        ]
        y = _start_pdf_page(pdf, 'Relatorio de Recebimentos de Fornecedores', generated_at, filters, page_number)

        pdf.setFont('Helvetica', 9)
        pdf.setFillColor(colors.HexColor('#1b2a41'))
        pdf.drawString(
            32,
            y,
            _pdf_text(
                f"Total de recebimentos: {total_receipts} | Itens previstos: {total_items} | "
                f"Recebimentos conferidos: {conferred_total}"
            ),
        )
        y -= 18

        for receipt in receipts:
            items = list(receipt.items.all())
            required_height = 76 + (len(items) * 12)
            if y < required_height + 42:
                _draw_pdf_footer(pdf, page_number)
                pdf.showPage()
                page_number += 1
                y = _start_pdf_page(pdf, 'Relatorio de Recebimentos de Fornecedores', generated_at, filters, page_number)

            school_label = receipt.school.name if receipt.school else 'Estoque Central'
            pdf.setFillColor(colors.HexColor('#f7f9fc'))
            pdf.rect(32, y - required_height + 8, 530, required_height, stroke=1, fill=1)
            pdf.setFillColor(colors.HexColor('#1b2a41'))
            pdf.setFont('Helvetica-Bold', 10)
            pdf.drawString(
                40,
                y - 8,
                _pdf_text(f"{receipt.expected_date} | {receipt.supplier.name} | {school_label} | {receipt.get_status_display()}"),
            )

            pdf.setFont('Helvetica', 8.5)
            sender = receipt.sender_signed_by or '-'
            receiver = receipt.receiver_signed_by or '-'
            pdf.drawString(40, y - 22, _pdf_text(f"Entregador: {sender[:40]}"))
            pdf.drawString(300, y - 22, _pdf_text(f"Recebedor: {receiver[:35]}"))

            pdf.setFont('Helvetica-Bold', 8)
            pdf.drawString(40, y - 36, _pdf_text('Item'))
            pdf.drawRightString(392, y - 36, _pdf_text('Prev.'))
            pdf.drawRightString(462, y - 36, _pdf_text('Receb.'))
            pdf.drawString(470, y - 36, _pdf_text('Obs.'))
            pdf.setFont('Helvetica', 8)
            line_y = y - 48

            for item in items:
                supply_name = (
                    item.supply.name
                    if item.supply
                    else (item.supply_created.name if item.supply_created else item.raw_name)
                ) or 'Item sem nome'
                received = '-' if item.received_quantity is None else f"{float(item.received_quantity):.2f} {item.unit}"
                pdf.drawString(40, line_y, _pdf_text(supply_name[:50]))
                pdf.drawRightString(392, line_y, _pdf_text(f"{float(item.expected_quantity):.2f} {item.unit}"))
                pdf.drawRightString(462, line_y, _pdf_text(received))
                pdf.drawString(470, line_y, _pdf_text((item.divergence_note or '-')[:14]))
                line_y -= 12

            y -= required_height + 8

        _draw_pdf_footer(pdf, page_number)
        pdf.save()
        return response


class DeliveryViewSet(viewsets.ModelViewSet):
    queryset = Delivery.objects.select_related('school', 'created_by').prefetch_related('items__supply').all().order_by('-created_at')
    serializer_class = DeliverySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        school = self.request.query_params.get('school')
        status_value = self.request.query_params.get('status')
        conference_enabled = self.request.query_params.get('conference_enabled')

        if school:
            queryset = queryset.filter(school_id=school)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if conference_enabled in ['true', 'false']:
            queryset = queryset.filter(conference_enabled=conference_enabled == 'true')
        return queryset

    def perform_destroy(self, instance):
        if instance.status != Delivery.Status.DRAFT:
            raise ValidationError('Apenas entregas em rascunho podem ser excluidas.')
        instance.delete()

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        delivery = self.get_object()
        if delivery.status != Delivery.Status.DRAFT:
            raise ValidationError('Somente entregas em rascunho podem ser enviadas.')

        items = list(delivery.items.select_related('supply').all())
        if not items:
            raise ValidationError('Adicione itens antes de enviar a entrega.')

        with transaction.atomic():
            for item in items:
                balance, _ = StockBalance.objects.select_for_update().get_or_create(supply=item.supply)
                if balance.quantity < item.planned_quantity:
                    raise ValidationError(
                        f"Saldo insuficiente para {item.supply.name}. Disponivel: {balance.quantity}"
                    )
                balance.quantity -= item.planned_quantity
                balance.save()
                StockMovement.objects.create(
                    supply=item.supply,
                    school=delivery.school,
                    type=StockMovement.Types.OUT,
                    quantity=item.planned_quantity,
                    movement_date=delivery.delivery_date,
                    note=f"Saida automatica da entrega {delivery.id} para {delivery.school.name}.",
                    created_by=request.user,
                )

            delivery.status = Delivery.Status.SENT
            delivery.conference_enabled = True
            delivery.sent_at = timezone.now()
            delivery.save(update_fields=['status', 'conference_enabled', 'sent_at', 'updated_at'])

        serializer = self.get_serializer(delivery)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def conference_link(self, request, pk=None):
        delivery = self.get_object()
        return Response({
            'slug': delivery.school.public_slug,
            'token': delivery.school.public_token,
            'delivery_id': str(delivery.id),
            'url': f"/public/delivery?slug={delivery.school.public_slug}&token={delivery.school.public_token}&delivery_id={delivery.id}",
            'api_url': f"/public/schools/{delivery.school.public_slug}/delivery/current/?token={delivery.school.public_token}&delivery_id={delivery.id}",
        })


class NotificationViewSet(viewsets.ModelViewSet):
    """ViewSet for listing and managing notifications."""
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.all().order_by('-created_at')

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        count = self.get_queryset().filter(is_read=False).count()
        return Response({'count': count})

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save(update_fields=['is_read'])
        return Response(NotificationSerializer(notification).data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        self.get_queryset().filter(is_read=False).update(is_read=True)
        return Response({'status': 'ok'})


class SchoolStockConfigViewSet(viewsets.ModelViewSet):
    """ViewSet for configuring stock limits per school."""
    queryset = SchoolStockBalance.objects.select_related('school', 'supply').all()
    serializer_class = SchoolStockBalanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        school = self.request.query_params.get('school')
        if school:
            queryset = queryset.filter(school_id=school)
        return queryset.order_by('supply__name')

    @action(detail=False, methods=['post'])
    def bulk_update_limits(self, request):
        """Update min_stock limits for multiple items at once."""
        items = request.data.get('items', [])
        if not items:
            return Response({'detail': 'Nenhum item informado.'}, status=400)

        payload = SchoolStockBulkLimitItemSerializer(data=items, many=True)
        payload.is_valid(raise_exception=True)

        validated_items = payload.validated_data
        balance_ids = [entry['id'] for entry in validated_items]
        balances = {
            balance.id: balance
            for balance in SchoolStockBalance.objects.filter(id__in=balance_ids)
        }
        missing_ids = [balance_id for balance_id in balance_ids if balance_id not in balances]
        if missing_ids:
            raise ValidationError({'detail': f'Itens de estoque não encontrados: {missing_ids}'})

        with transaction.atomic():
            for entry in validated_items:
                balance = balances[entry['id']]
                balance.min_stock = entry['min_stock']
                balance.save(update_fields=['min_stock'])

        updated = len(validated_items)
        return Response({'updated': updated})

    @action(detail=True, methods=['patch'])
    def update_limit(self, request, pk=None):
        """Update min_stock limit for a single item."""
        balance = self.get_object()
        payload = SchoolStockLimitUpdateSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        balance.min_stock = payload.validated_data['min_stock']
        balance.save(update_fields=['min_stock'])
        return Response(SchoolStockBalanceSerializer(balance).data)
